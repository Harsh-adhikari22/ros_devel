import openpyxl
import matplotlib.pyplot as plt
loc=r'/Users/harshadhikari/Desktop/RiceProductionAreaWise(Vietnam).xlsx'
wb=openpyxl.load_workbook(loc)
sheet=wb.active
names=[]
values=[]
for i in range(2,50):
    names.append(sheet.cell(i,1).value)
    values.append(sheet.cell(i,2).value)
plt.xkcd()
plt.yticks(ticks=[i/2 for i in range(0,14)])
plt.plot(names,values,color='b',markerfacecolor='k',marker='o')
plt.tick_params(axis='x', which='major', labelrotation=90)
plt.fill_between(names,values,color='b',alpha=0.4)
plt.xlabel("Cities in Vietnam")
plt.ylabel("Percentage Yield of Rice")
plt.grid(True)
plt.show()
