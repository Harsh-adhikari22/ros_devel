import openpyxl
import matplotlib.pyplot as plt
loc=r'/Users/harshadhikari/Desktop/MekongWeather.xlsx'
wb=openpyxl.load_workbook(loc)
sheet=wb.active
months=[]
PP=[]
for i in range(2,14):
    months.append(sheet.cell(i,1).value)
    PP.append(sheet.cell(i,5).value)
plt.grid(True)
plt.plot(months,PP,color='Black',marker='o',label='Precipitation per month')
plt.fill_between(months,PP,y2=-10,color="#000088",alpha=0.5)
plt.xlabel("Months")
plt.ylabel("Rainfall (in mm)")
plt.yticks(ticks=[i*12 for i in range(12)])
plt.ylim(-10,120)
plt.title("Rainfall")
for i in range(12):
    plt.annotate(text=PP[i],xy=(months[i],PP[i]-6))
plt.show()

